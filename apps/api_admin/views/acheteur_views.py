from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from drf_spectacular.utils import extend_schema

from apps.accounts.permissions import IsSuperAdmin
from apps.accounts.models import Acheteur
from apps.payments.models import Voucher, ProgrammeVoucher


# ── Helpers ──────────────────────────────────────────────────────

def _programme_data(p, with_vouchers=False):
    d = {
        'id':             p.pk,
        'nom':            p.nom,
        'code_programme': p.code_programme,
        'type_programme': p.type_programme,
        'type_label':     p.get_type_programme_display(),
        'description':    p.description,
        'logo':           p.logo.url if p.logo else None,
        'contact_nom':    p.contact_nom,
        'contact_email':  p.contact_email,
        'contact_tel':    p.contact_tel,
        'budget_total':   str(p.budget_total) if p.budget_total is not None else None,
        'budget_utilise': str(p.budget_utilise),
        'budget_restant': str(p.budget_restant) if p.budget_restant is not None else None,
        'is_active':      p.is_active,
        'est_en_cours':   p.est_en_cours,
        'date_debut':     p.date_debut.isoformat(),
        'date_fin':       p.date_fin.isoformat(),
        'nb_vouchers':    p.vouchers.count(),
        'created_at':     p.created_at.isoformat(),
    }
    if with_vouchers:
        d['vouchers'] = [_voucher_data(v) for v in p.vouchers.select_related('beneficiaire__user').all()]
    return d


def _voucher_data(v):
    return {
        'id':                   v.pk,
        'code':                 v.code,
        'programme_id':         v.programme_id,
        'programme_nom':        v.programme.nom,
        'beneficiaire_id':      v.beneficiaire_id,
        'beneficiaire_nom':     v.beneficiaire.user.get_full_name() if v.beneficiaire else None,
        'beneficiaire_email':   v.beneficiaire.user.email if v.beneficiaire else None,
        'type_valeur':          v.type_valeur,
        'type_label':           v.get_type_valeur_display(),
        'valeur':               str(v.valeur),
        'montant_max':          str(v.montant_max) if v.montant_max is not None else None,
        'montant_commande_min': str(v.montant_commande_min),
        'statut':               v.statut,
        'statut_label':         v.get_statut_display(),
        'date_expiration':      v.date_expiration.isoformat(),
        'date_utilisation':     v.date_utilisation.isoformat() if v.date_utilisation else None,
        'est_valide':           v.est_valide,
        'cree_par':             v.cree_par.get_full_name() if v.cree_par else None,
        'created_at':           v.created_at.isoformat(),
    }


# ── Acheteurs ────────────────────────────────────────────────────

@extend_schema(operation_id='admin_acheteurs_list', tags=['Admin — Acheteurs'])
@api_view(['GET'])
@permission_classes([IsSuperAdmin])
def acheteurs_list(request):
    qs   = Acheteur.objects.select_related('user').order_by('-created_at')
    data = [
        {
            'id':               a.pk,
            'full_name':        a.user.get_full_name(),
            'email':            a.user.email,
            'telephone':        a.user.telephone,
            'type_acheteur':    a.type_acheteur,
            'type_label':       a.get_type_acheteur_display(),
            'nom_organisation': a.nom_organisation,
            'departement':      a.departement,
            'is_active':        a.user.is_active,
            'total_commandes':  a.total_commandes,
            'total_depense':    str(a.total_depense),
            'created_at':       a.created_at.isoformat(),
        }
        for a in qs
    ]
    return Response({'success': True, 'data': data})


@extend_schema(operation_id='admin_acheteur_detail', tags=['Admin — Acheteurs'])
@api_view(['GET'])
@permission_classes([IsSuperAdmin])
def acheteur_detail(request, pk):
    a = get_object_or_404(Acheteur, pk=pk)
    return Response({
        'success': True,
        'data': {
            'id':               a.pk,
            'full_name':        a.user.get_full_name(),
            'email':            a.user.email,
            'telephone':        a.user.telephone,
            'type_acheteur':    a.type_acheteur,
            'type_label':       a.get_type_acheteur_display(),
            'nom_organisation': a.nom_organisation,
            'adresse':          a.adresse,
            'ville':            a.ville,
            'departement':      a.departement,
            'is_active':        a.user.is_active,
            'total_commandes':  a.total_commandes,
            'total_depense':    str(a.total_depense),
        }
    })


# ── Programmes Voucher ───────────────────────────────────────────

@extend_schema(operation_id='admin_programmes_list', tags=['Admin — Vouchers'])
@api_view(['GET', 'POST'])
@permission_classes([IsSuperAdmin])
def programmes_list(request):
    """
    GET  — liste tous les programmes.
    POST — crée un nouveau programme.
    Body POST : { nom, code_programme, type_programme, description?,
                  contact_nom?, contact_email?, contact_tel?,
                  budget_total?, date_debut, date_fin }
    """
    if request.method == 'POST':
        nom           = request.data.get('nom', '').strip()
        code          = request.data.get('code_programme', '').strip().upper()
        type_prog     = request.data.get('type_programme', '').strip()
        date_debut    = request.data.get('date_debut')
        date_fin      = request.data.get('date_fin')

        errors = {}
        if not nom:
            errors['nom'] = 'Le nom est requis.'
        if not code:
            errors['code_programme'] = 'Le code programme est requis.'
        elif ProgrammeVoucher.objects.filter(code_programme=code).exists():
            errors['code_programme'] = f"Le code '{code}' est déjà utilisé."
        if not type_prog or type_prog not in dict(ProgrammeVoucher.TypeProgramme.choices):
            errors['type_programme'] = (
                f"Type invalide. Choix : {list(dict(ProgrammeVoucher.TypeProgramme.choices).keys())}"
            )
        if not date_debut:
            errors['date_debut'] = 'La date de début est requise.'
        if not date_fin:
            errors['date_fin'] = 'La date de fin est requise.'

        if errors:
            return Response({'success': False, 'error': errors},
                            status=status.HTTP_400_BAD_REQUEST)

        # Vérification cohérence des dates
        from datetime import date as date_type
        try:
            from django.utils.dateparse import parse_date
            d_debut = parse_date(date_debut)
            d_fin   = parse_date(date_fin)
        except Exception:
            return Response(
                {'success': False, 'error': 'Format de date invalide (YYYY-MM-DD).'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if d_debut and d_fin and d_fin < d_debut:
            return Response(
                {'success': False, 'error': 'La date de fin doit être après la date de début.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        budget_raw = request.data.get('budget_total')
        try:
            from decimal import Decimal
            budget_total = Decimal(str(budget_raw)) if budget_raw not in (None, '') else None
        except (ValueError, TypeError, Exception):
            return Response(
                {'success': False, 'error': 'budget_total doit être un nombre.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        programme = ProgrammeVoucher.objects.create(
            nom             = nom,
            code_programme  = code,
            type_programme  = type_prog,
            description     = request.data.get('description', '').strip(),
            contact_nom     = request.data.get('contact_nom', '').strip(),
            contact_email   = request.data.get('contact_email', '').strip(),
            contact_tel     = request.data.get('contact_tel', '').strip(),
            budget_total    = budget_total,
            date_debut      = d_debut,
            date_fin        = d_fin,
            is_active       = True,
        )
        return Response(
            {'success': True, 'data': _programme_data(programme)},
            status=status.HTTP_201_CREATED,
        )

    # GET
    progs = ProgrammeVoucher.objects.all().order_by('-created_at')
    return Response({'success': True, 'data': [_programme_data(p) for p in progs]})


@extend_schema(operation_id='admin_programme_detail', tags=['Admin — Vouchers'])
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsSuperAdmin])
def programme_detail(request, pk):
    """
    GET    — détail du programme avec la liste de ses vouchers.
    PATCH  — modifier nom, description, dates, budget, is_active, contact.
    DELETE — supprimer uniquement si aucun voucher n'a été utilisé.
    """
    p = get_object_or_404(ProgrammeVoucher, pk=pk)

    if request.method == 'DELETE':
        # Refuser si des vouchers ont déjà été utilisés
        if p.vouchers.filter(statut=Voucher.Statut.UTILISE).exists():
            return Response(
                {'success': False,
                 'error': "Impossible de supprimer : des vouchers de ce programme ont déjà été utilisés."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        p.vouchers.all().delete()
        p.delete()
        return Response({'success': True, 'data': {'id': pk}})

    if request.method == 'PATCH':
        for field in ('nom', 'description', 'contact_nom', 'contact_email', 'contact_tel'):
            if field in request.data:
                setattr(p, field, request.data[field].strip() if isinstance(request.data[field], str) else request.data[field])

        if 'is_active' in request.data:
            val = request.data['is_active']
            p.is_active = val if isinstance(val, bool) else str(val).lower() != 'false'

        if 'budget_total' in request.data:
            raw = request.data['budget_total']
            try:
                from decimal import Decimal
                p.budget_total = Decimal(str(raw)) if raw not in (None, '') else None
            except (ValueError, TypeError, Exception):
                return Response(
                    {'success': False, 'error': 'budget_total doit être un nombre.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        from django.utils.dateparse import parse_date
        if 'date_debut' in request.data:
            p.date_debut = parse_date(request.data['date_debut'])
        if 'date_fin' in request.data:
            p.date_fin = parse_date(request.data['date_fin'])

        if p.date_fin and p.date_debut and p.date_fin < p.date_debut:
            return Response(
                {'success': False, 'error': 'La date de fin doit être après la date de début.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        p.save()
        return Response({'success': True, 'data': _programme_data(p)})

    # GET — inclure les vouchers du programme
    return Response({'success': True, 'data': _programme_data(p, with_vouchers=True)})


# ── Vouchers ─────────────────────────────────────────────────────

@extend_schema(operation_id='admin_vouchers_list', tags=['Admin — Vouchers'])
@api_view(['GET', 'POST'])
@permission_classes([IsSuperAdmin])
def vouchers_list(request):
    """
    GET  — liste tous les vouchers (filtre par programme_id, statut).
    POST — crée un voucher individuel sur un programme.
    Body POST : { programme_id, type_valeur, valeur, date_expiration,
                  beneficiaire_id?, montant_max?, montant_commande_min?,
                  categories_autorisees? }
    """
    if request.method == 'POST':
        return _creer_voucher(request)

    # GET avec filtres optionnels
    qs = Voucher.objects.select_related(
        'programme', 'beneficiaire__user', 'cree_par'
    ).order_by('-created_at')

    programme_id = request.query_params.get('programme_id')
    statut_f     = request.query_params.get('statut')
    search_q     = request.query_params.get('search', '').strip()
    if programme_id:
        qs = qs.filter(programme_id=programme_id)
    if statut_f:
        qs = qs.filter(statut=statut_f)
    if search_q:
        from django.db.models import Q
        qs = qs.filter(
            Q(code__icontains=search_q) |
            Q(beneficiaire__user__email__icontains=search_q) |
            Q(beneficiaire__user__first_name__icontains=search_q) |
            Q(beneficiaire__user__last_name__icontains=search_q)
        )

    return Response({'success': True, 'data': [_voucher_data(v) for v in qs]})


@extend_schema(operation_id='admin_voucher_detail', tags=['Admin — Vouchers'])
@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsSuperAdmin])
def voucher_detail(request, pk):
    """
    GET    — détail d'un voucher.
    PATCH  — modifier statut, date_expiration, beneficiaire_id.
             Un voucher UTILISE ne peut pas être modifié.
    DELETE — annuler (met statut=annule). Pas de suppression physique si utilisé.
    """
    v = get_object_or_404(
        Voucher.objects.select_related('programme', 'beneficiaire__user', 'cree_par'),
        pk=pk,
    )

    if request.method == 'DELETE':
        if v.statut == Voucher.Statut.UTILISE:
            return Response(
                {'success': False, 'error': "Un voucher déjà utilisé ne peut pas être annulé."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        v.statut = Voucher.Statut.ANNULE
        v.save(update_fields=['statut'])
        return Response({'success': True, 'data': _voucher_data(v)})

    if request.method == 'PATCH':
        if v.statut == Voucher.Statut.UTILISE:
            return Response(
                {'success': False, 'error': "Un voucher déjà utilisé ne peut pas être modifié."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        STATUTS_VALIDES = [
            Voucher.Statut.ACTIF,
            Voucher.Statut.SUSPENDU,
            Voucher.Statut.ANNULE,
            Voucher.Statut.EXPIRE,
        ]
        if 'statut' in request.data:
            nouveau = request.data['statut']
            if nouveau not in [s.value for s in STATUTS_VALIDES]:
                return Response(
                    {'success': False, 'error': f"Statut invalide. Choix : {[s.value for s in STATUTS_VALIDES]}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            v.statut = nouveau

        if 'date_expiration' in request.data:
            from django.utils.dateparse import parse_date
            d = parse_date(request.data['date_expiration'])
            if not d:
                return Response(
                    {'success': False, 'error': 'Format de date invalide (YYYY-MM-DD).'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            v.date_expiration = d

        if 'beneficiaire_id' in request.data:
            bid = request.data['beneficiaire_id']
            if bid in (None, ''):
                v.beneficiaire = None
            else:
                try:
                    v.beneficiaire = Acheteur.objects.get(pk=int(bid))
                except (Acheteur.DoesNotExist, ValueError, TypeError):
                    return Response(
                        {'success': False, 'error': f"Acheteur #{bid} introuvable."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

        v.save()
        return Response({'success': True, 'data': _voucher_data(v)})

    # GET
    return Response({'success': True, 'data': _voucher_data(v)})


@extend_schema(operation_id='admin_vouchers_bulk_create', tags=['Admin — Vouchers'])
@api_view(['POST'])
@permission_classes([IsSuperAdmin])
def vouchers_bulk_create(request):
    """
    Créer N vouchers identiques sur un programme.
    Body : { programme_id, nombre, type_valeur, valeur, date_expiration,
             beneficiaire_ids?  (liste d'Acheteur pk — si fournie, doit avoir exactement `nombre` éléments),
             montant_max?, montant_commande_min?, categories_autorisees? }
    """
    programme_id    = request.data.get('programme_id')
    nombre_raw      = request.data.get('nombre', 1)
    beneficiaire_ids = request.data.get('beneficiaire_ids', [])

    try:
        nombre = int(nombre_raw)
        if nombre < 1 or nombre > 500:
            raise ValueError
    except (ValueError, TypeError):
        return Response(
            {'success': False, 'error': 'nombre doit être un entier entre 1 et 500.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    programme = get_object_or_404(ProgrammeVoucher, pk=programme_id)

    if beneficiaire_ids and len(beneficiaire_ids) != nombre:
        return Response(
            {'success': False,
             'error': f"beneficiaire_ids doit contenir exactement {nombre} éléments (un par voucher)."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Valider les bénéficiaires si fournis
    beneficiaires = []
    if beneficiaire_ids:
        for bid in beneficiaire_ids:
            try:
                beneficiaires.append(Acheteur.objects.get(pk=int(bid)))
            except (Acheteur.DoesNotExist, ValueError, TypeError):
                return Response(
                    {'success': False, 'error': f"Acheteur #{bid} introuvable."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

    # Valider les champs communs
    erreur, commun = _valider_champs_voucher(request.data)
    if erreur:
        return Response({'success': False, 'error': erreur},
                        status=status.HTTP_400_BAD_REQUEST)

    # Créer les N vouchers
    categories = commun.pop('_categories', [])
    crees = []
    for i in range(nombre):
        v = Voucher(
            programme   = programme,
            beneficiaire= beneficiaires[i] if beneficiaires else None,
            cree_par    = request.user,
            **commun,
        )
        v.save()  # génère le code unique dans save()
        if categories:
            v.categories_autorisees.set(categories)
        crees.append(v)

    # Recharger pour avoir les relations
    ids = [v.pk for v in crees]
    crees_qs = Voucher.objects.filter(pk__in=ids).select_related(
        'programme', 'beneficiaire__user', 'cree_par'
    )
    crees_list = list(crees_qs)

    # Envoyer un email à chaque bénéficiaire
    from apps.emails.utils import email_voucher_cree
    for v in crees_list:
        try:
            email_voucher_cree(v)
        except Exception:
            import logging
            logging.getLogger(__name__).exception("Erreur envoi email voucher %s", v.code)

    return Response(
        {
            'success': True,
            'data': {
                'nb_crees':  nombre,
                'programme': programme.nom,
                'vouchers':  [_voucher_data(v) for v in crees_list],
            }
        },
        status=status.HTTP_201_CREATED,
    )


# ── Helpers internes ─────────────────────────────────────────────

def _valider_champs_voucher(data):
    """
    Valide et normalise les champs communs à la création d'un voucher.
    Retourne (erreur_dict_ou_str | None, champs_dict).
    """
    from django.utils.dateparse import parse_date
    from apps.catalog.models import Categorie

    type_valeur = data.get('type_valeur', '').strip()
    valeur_raw  = data.get('valeur')
    exp_raw     = data.get('date_expiration')

    errors = {}
    if type_valeur not in (Voucher.TypeValeur.FIXE, Voucher.TypeValeur.POURCENTAGE):
        errors['type_valeur'] = "Valeur invalide. Choix : 'fixe', 'pourcent'."
    try:
        valeur = float(valeur_raw)
        if valeur <= 0:
            raise ValueError
    except (ValueError, TypeError):
        errors['valeur'] = 'valeur doit être un nombre positif.'
    if not exp_raw:
        errors['date_expiration'] = 'La date d\'expiration est requise.'
    else:
        date_exp = parse_date(str(exp_raw))
        if not date_exp:
            errors['date_expiration'] = 'Format invalide (YYYY-MM-DD).'

    if errors:
        return errors, {}

    montant_max_raw = data.get('montant_max')
    try:
        montant_max = float(montant_max_raw) if montant_max_raw not in (None, '') else None
    except (ValueError, TypeError):
        return {'montant_max': 'montant_max doit être un nombre.'}, {}

    montant_min_raw = data.get('montant_commande_min', 0)
    try:
        montant_min = float(montant_min_raw)
    except (ValueError, TypeError):
        return {'montant_commande_min': 'montant_commande_min doit être un nombre.'}, {}

    # Catégories autorisées (optionnel)
    cat_ids = data.get('categories_autorisees', [])
    categories = []
    if cat_ids:
        try:
            categories = list(Categorie.objects.filter(pk__in=[int(c) for c in cat_ids]))
        except (ValueError, TypeError):
            return {'categories_autorisees': 'Liste d\'IDs de catégories invalide.'}, {}

    champs = {
        'type_valeur':          type_valeur,
        'valeur':               valeur,
        'montant_max':          montant_max,
        'montant_commande_min': montant_min,
        'date_expiration':      parse_date(str(exp_raw)),
        '_categories':          categories,  # set M2M séparément
    }
    return None, champs


def _creer_voucher(request):
    """Crée un seul voucher via POST /api/admin/vouchers/."""
    programme_id   = request.data.get('programme_id')
    beneficiaire_id = request.data.get('beneficiaire_id')

    if not programme_id:
        return Response(
            {'success': False, 'error': 'programme_id est requis.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    programme = get_object_or_404(ProgrammeVoucher, pk=programme_id)

    beneficiaire = None
    if beneficiaire_id not in (None, ''):
        try:
            beneficiaire = Acheteur.objects.get(pk=int(beneficiaire_id))
        except (Acheteur.DoesNotExist, ValueError, TypeError):
            return Response(
                {'success': False, 'error': f"Acheteur #{beneficiaire_id} introuvable."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    erreur, champs = _valider_champs_voucher(request.data)
    if erreur:
        return Response({'success': False, 'error': erreur},
                        status=status.HTTP_400_BAD_REQUEST)

    categories = champs.pop('_categories', [])
    v = Voucher(
        programme    = programme,
        beneficiaire = beneficiaire,
        cree_par     = request.user,
        **champs,
    )
    v.save()
    if categories:
        v.categories_autorisees.set(categories)

    # Recharger pour les relations
    v.refresh_from_db()
    v = Voucher.objects.select_related('programme', 'beneficiaire__user', 'cree_par').get(pk=v.pk)
    return Response(
        {'success': True, 'data': _voucher_data(v)},
        status=status.HTTP_201_CREATED,
    )


# ── Import Excel ─────────────────────────────────────────────────

@extend_schema(operation_id='admin_vouchers_template_excel', tags=['Admin — Vouchers'])
@api_view(['GET'])
@permission_classes([IsSuperAdmin])
def vouchers_template_excel(request):
    """Télécharger le modèle Excel pour l'import de bénéficiaires."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Bénéficiaires'

    # En-têtes
    headers = ['email', 'nom_complet']
    header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Exemples
    examples = [
        ['jean.duval@example.com', 'Jean Duval'],
        ['marie.pierre@example.com', 'Marie Pierre'],
    ]
    example_font = Font(color='555555', italic=True)
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.border = thin_border

    # Largeurs de colonnes
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 28

    # Feuille d'instructions
    ws_info = wb.create_sheet(title='Instructions')
    instructions = [
        ('Colonne', 'Description', 'Obligatoire'),
        ('email',       'Adresse e-mail du compte Maket Peyizan',  'Oui'),
        ('nom_complet', 'Nom affiché (informatif seulement)',        'Non'),
    ]
    for r_idx, row in enumerate(instructions, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_info.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
    ws_info.column_dimensions['A'].width = 16
    ws_info.column_dimensions['B'].width = 44
    ws_info.column_dimensions['C'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="template_beneficiaires.xlsx"'
    return response


@extend_schema(operation_id='admin_vouchers_import_excel', tags=['Admin — Vouchers'])
@api_view(['POST'])
@permission_classes([IsSuperAdmin])
def vouchers_import_excel(request):
    """
    Importer des bénéficiaires depuis un fichier Excel et créer un voucher par ligne.
    Body (multipart/form-data) :
      file            — fichier .xlsx
      programme_id    — int
      type_valeur     — 'fixe' | 'pourcent'
      valeur          — nombre
      date_expiration — YYYY-MM-DD
      montant_commande_min? — nombre
    """
    import io
    import openpyxl

    fichier = request.FILES.get('file')
    if not fichier:
        return Response({'success': False, 'error': 'Aucun fichier fourni.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Lire le fichier Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(fichier.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return Response({'success': False, 'error': 'Fichier Excel invalide ou illisible.'},
                        status=status.HTTP_400_BAD_REQUEST)

    if len(rows) < 2:
        return Response({'success': False, 'error': 'Le fichier ne contient aucune ligne de données (hors en-tête).'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Trouver la colonne "email" dans la première ligne (en-tête)
    header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
    try:
        email_col = header.index('email')
    except ValueError:
        return Response({'success': False, 'error': "Colonne 'email' introuvable dans le fichier."},
                        status=status.HTTP_400_BAD_REQUEST)

    # Extraire les e-mails
    emails = []
    for row_num, row in enumerate(rows[1:], start=2):
        if email_col >= len(row):
            continue
        email_val = row[email_col]
        if email_val is None:
            continue
        email_str = str(email_val).strip().lower()
        if email_str:
            emails.append((row_num, email_str))

    if not emails:
        return Response({'success': False, 'error': 'Aucun e-mail trouvé dans le fichier.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Résoudre les e-mails en Acheteur
    from apps.accounts.models import CustomUser
    beneficiaires = []
    erreurs = []
    for row_num, email in emails:
        try:
            user = CustomUser.objects.get(email__iexact=email)
            acheteur = Acheteur.objects.get(user=user)
            beneficiaires.append(acheteur)
        except CustomUser.DoesNotExist:
            erreurs.append(f"Ligne {row_num} : aucun compte avec l'email « {email} ».")
        except Acheteur.DoesNotExist:
            erreurs.append(f"Ligne {row_num} : l'utilisateur « {email} » n'a pas de profil acheteur.")

    if erreurs:
        return Response({'success': False, 'error': erreurs},
                        status=status.HTTP_400_BAD_REQUEST)

    if not beneficiaires:
        return Response({'success': False, 'error': 'Aucun bénéficiaire valide trouvé.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Valider le programme
    programme_id = request.data.get('programme_id')
    if not programme_id:
        return Response({'success': False, 'error': 'programme_id est requis.'},
                        status=status.HTTP_400_BAD_REQUEST)
    programme = get_object_or_404(ProgrammeVoucher, pk=programme_id)

    # Valider les champs du voucher
    erreur, commun = _valider_champs_voucher(request.data)
    if erreur:
        return Response({'success': False, 'error': erreur},
                        status=status.HTTP_400_BAD_REQUEST)

    # Créer les vouchers
    categories = commun.pop('_categories', [])
    crees = []
    for beneficiaire in beneficiaires:
        v = Voucher(
            programme    = programme,
            beneficiaire = beneficiaire,
            cree_par     = request.user,
            **commun,
        )
        v.save()
        if categories:
            v.categories_autorisees.set(categories)
        crees.append(v)

    ids = [v.pk for v in crees]
    crees_qs = Voucher.objects.filter(pk__in=ids).select_related(
        'programme', 'beneficiaire__user', 'cree_par'
    )
    crees_list = list(crees_qs)

    # Envoyer un email à chaque bénéficiaire
    from apps.emails.utils import email_voucher_cree
    import logging as _logging
    for v in crees_list:
        try:
            email_voucher_cree(v)
        except Exception:
            _logging.getLogger(__name__).exception("Erreur envoi email voucher %s", v.code)

    return Response(
        {
            'success': True,
            'data': {
                'nb_crees':  len(crees),
                'programme': programme.nom,
                'vouchers':  [_voucher_data(v) for v in crees_list],
            }
        },
        status=status.HTTP_201_CREATED,
    )


# ── Adresses ─────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsSuperAdmin])
def adresses_list_admin(request):
    from apps.accounts.models import Adresse
    from apps.accounts.serializers import AdresseSerializer
    qs = Adresse.objects.select_related('user').order_by('-created_at')
    return Response({
        'success': True,
        'data': AdresseSerializer(qs, many=True).data
    })
