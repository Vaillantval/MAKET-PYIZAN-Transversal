from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.utils import timezone
from django.db.models import Sum, Count, Avg, Q
from datetime import timedelta, date, datetime
from io import BytesIO


@method_decorator(staff_member_required, name='dispatch')
class DashboardView(TemplateView):
    template_name = 'analytics/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            'kpis':                    self._get_kpis(),
            'ventes_par_jour':         self._get_ventes_par_jour(),
            'ventes_par_mois':         self._get_ventes_par_mois(),
            'commandes_par_statut':    self._get_commandes_par_statut(),
            'paiements_par_type':      self._get_paiements_par_type(),
            'top_produits':            self._get_top_produits(),
            'top_producteurs':         self._get_top_producteurs(),
            'top_acheteurs':           self._get_top_acheteurs(),
            'producteurs_par_dept':    self._get_producteurs_par_departement(),
            'ventes_par_categorie':    self._get_ventes_par_categorie(),
            'alertes_stock':           self._get_alertes_stock(),
            'collectes_a_venir':       self._get_collectes_a_venir(),
            'collectes_en_retard':     self._get_collectes_en_retard(),
            'dernieres_commandes':     self._get_dernieres_commandes(),
            'derniers_producteurs':    self._get_derniers_producteurs(),
            'paiements_en_attente':    self._get_paiements_en_attente(),
            'today':                   timezone.now(),
        })
        return ctx

    def _get_kpis(self):
        from apps.accounts.models import CustomUser, Producteur, Acheteur
        from apps.catalog.models import Produit
        from apps.orders.models import Commande
        from apps.payments.models import Paiement
        from apps.stock.models import AlerteStock

        today = timezone.now().date()
        debut_mois = today.replace(day=1)

        total_commandes = Commande.objects.count()
        commandes_mois  = Commande.objects.filter(created_at__date__gte=debut_mois).count()

        ca_total = Commande.objects.filter(statut_paiement='paye').aggregate(total=Sum('total'))['total'] or 0
        ca_mois  = Commande.objects.filter(statut_paiement='paye', created_at__date__gte=debut_mois).aggregate(total=Sum('total'))['total'] or 0

        return {
            'total_utilisateurs':   CustomUser.objects.filter(is_active=True).count(),
            'total_producteurs':    Producteur.objects.filter(statut='actif').count(),
            'total_acheteurs':      Acheteur.objects.count(),
            'total_produits':       Produit.objects.filter(is_active=True).count(),
            'total_commandes':      total_commandes,
            'commandes_mois':       commandes_mois,
            'ca_total':             ca_total,
            'ca_mois':              ca_mois,
            'alertes_nouvelles':    AlerteStock.objects.filter(statut='nouvelle').count(),
            'paiements_en_attente': Paiement.objects.filter(statut__in=['initie', 'en_attente', 'soumis']).count(),
        }

    def _get_ventes_par_jour(self):
        from apps.orders.models import Commande
        today = timezone.now().date()
        data  = []
        for i in range(29, -1, -1):
            jour = today - timedelta(days=i)
            ca   = Commande.objects.filter(created_at__date=jour, statut_paiement='paye').aggregate(total=Sum('total'))['total'] or 0
            data.append({'date': jour.strftime('%d/%m'), 'ca': float(ca)})
        return data

    def _get_ventes_par_mois(self):
        from apps.orders.models import Commande
        today = timezone.now().date()
        data  = []
        for i in range(11, -1, -1):
            mois = (today.replace(day=1) - timedelta(days=i * 30)).replace(day=1)
            ca   = Commande.objects.filter(
                created_at__year=mois.year,
                created_at__month=mois.month,
                statut_paiement='paye'
            ).aggregate(total=Sum('total'))['total'] or 0
            data.append({'mois': mois.strftime('%b %Y'), 'ca': float(ca)})
        return data

    def _get_commandes_par_statut(self):
        from apps.orders.models import Commande
        statuts = Commande.objects.values('statut').annotate(count=Count('id')).order_by()
        return [{'statut': s['statut'], 'count': s['count']} for s in statuts]

    def _get_paiements_par_type(self):
        from apps.payments.models import Paiement
        types = Paiement.objects.filter(statut='confirme').values('type_paiement').annotate(
            count=Count('id'), total=Sum('montant')
        ).order_by('-total')
        return list(types)

    def _get_top_produits(self):
        from apps.orders.models import CommandeDetail
        return CommandeDetail.objects.values(
            'produit__nom', 'produit__slug'
        ).annotate(
            total_vendu=Sum('quantite'),
            ca=Sum('sous_total')
        ).order_by('-ca')[:10]

    def _get_top_producteurs(self):
        from apps.accounts.models import Producteur
        return Producteur.objects.annotate(
            nb_cmd=Count('commandes_recues'),
            ca=Sum('commandes_recues__total', filter=Q(commandes_recues__statut_paiement='paye'))
        ).order_by('-ca')[:10]

    def _get_top_acheteurs(self):
        from apps.accounts.models import Acheteur
        return Acheteur.objects.annotate(
            nb_cmd=Count('commandes'),
            depense=Sum('commandes__total', filter=Q(commandes__statut_paiement='paye'))
        ).order_by('-depense')[:10]

    def _get_producteurs_par_departement(self):
        from apps.accounts.models import Producteur
        return list(Producteur.objects.filter(statut='actif').values('departement').annotate(count=Count('id')).order_by('-count'))

    def _get_ventes_par_categorie(self):
        from apps.orders.models import CommandeDetail
        return list(CommandeDetail.objects.values('produit__categorie__nom').annotate(
            ca=Sum('sous_total'), nb=Count('id')
        ).order_by('-ca')[:8])

    def _get_alertes_stock(self):
        from apps.stock.models import AlerteStock
        return AlerteStock.objects.filter(
            statut__in=['nouvelle', 'vue']
        ).select_related('produit', 'produit__producteur__user').order_by('-created_at')[:15]

    def _get_collectes_a_venir(self):
        from apps.collectes.models import Collecte
        today = timezone.now().date()
        limit = today + timedelta(days=14)
        return Collecte.objects.filter(
            statut='planifiee',
            date_planifiee__range=[today, limit]
        ).select_related('zone', 'point_collecte', 'collecteur').order_by('date_planifiee')[:10]

    def _get_collectes_en_retard(self):
        from apps.collectes.models import Collecte
        today = timezone.now().date()
        return Collecte.objects.filter(
            statut__in=['planifiee', 'en_cours'],
            date_planifiee__lt=today
        ).select_related('zone', 'collecteur').order_by('date_planifiee')

    def _get_dernieres_commandes(self):
        from apps.orders.models import Commande
        return Commande.objects.select_related(
            'acheteur__user', 'producteur__user'
        ).order_by('-created_at')[:15]

    def _get_derniers_producteurs(self):
        from apps.accounts.models import Producteur
        return Producteur.objects.filter(
            statut='en_attente'
        ).select_related('user').order_by('-created_at')[:10]

    def _get_paiements_en_attente(self):
        from apps.payments.models import Paiement
        return Paiement.objects.filter(
            statut__in=['soumis', 'en_attente']
        ).select_related(
            'commande__acheteur__user', 'effectue_par'
        ).order_by('-created_at')[:15]


@method_decorator(staff_member_required, name='dispatch')
class ExportDashboardView(TemplateView):
    """View for exporting dashboard data"""
    template_name = 'analytics/export_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = timezone.now().date()
        ctx.update({
            'today': today,
            'default_start': today - timedelta(days=30),
            'default_end': today,
        })
        return ctx

    def post(self, request, *args, **kwargs):
        """Handle export request"""
        export_format = request.POST.get('format', 'pdf')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')

        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            today = timezone.now().date()
            start_date = today - timedelta(days=30)
            end_date = today

        # Generate report data
        from .report_generators import ReportDataGenerator
        gen = ReportDataGenerator(start_date, end_date)
        
        report_data = {
            'kpis': gen.get_kpis(),
            'daily_sales': gen.get_daily_sales(),
            'monthly_sales': gen.get_monthly_sales(),
            'orders_by_status': gen.get_orders_by_status(),
            'payments_by_type': gen.get_payments_by_type(),
            'top_products': gen.get_top_products(),
            'top_producers': gen.get_top_producers(),
            'top_buyers': gen.get_top_buyers(),
            'sales_by_category': list(gen.get_sales_by_category()),
        }

        if export_format == 'pdf':
            return self._export_pdf(report_data, start_date, end_date)
        elif export_format == 'csv':
            return self._export_csv(report_data, start_date, end_date)
        elif export_format == 'xlsx':
            return self._export_xlsx(report_data, start_date, end_date)
        else:
            return self._export_pdf(report_data, start_date, end_date)

    def _export_pdf(self, report_data, start_date, end_date):
        """Export as PDF"""
        from django.http import HttpResponse
        from .report_generators import PDFReportGenerator

        try:
            pdf_gen = PDFReportGenerator(report_data)
            pdf_buffer = pdf_gen.generate()
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            filename = f"rapport_maket_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError as e:
            from django.http import HttpResponseServerError
            return HttpResponseServerError(f'Erreur: reportlab n\'est pas installé. {str(e)}')
        except Exception as e:
            from django.http import HttpResponseServerError
            return HttpResponseServerError(f'Erreur lors de la génération du PDF: {str(e)}')


    def _export_csv(self, report_data, start_date, end_date):
        """Export as CSV"""
        from django.http import HttpResponse
        from .report_generators import CSVReportGenerator

        csv_gen = CSVReportGenerator(report_data)
        csv_buffer = csv_gen.generate()

        # UTF-8 BOM pour ouverture correcte des accents dans Excel
        response = HttpResponse(
            '\ufeff' + csv_buffer.getvalue(),
            content_type='text/csv; charset=utf-8-sig'
        )
        filename = f"rapport_maket_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _export_xlsx(self, report_data, start_date, end_date):
        """Export as Excel"""
        from django.http import HttpResponse, HttpResponseServerError
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return self._export_csv(report_data, start_date, end_date)

        try:
            wb = openpyxl.Workbook()
            
            # KPIs Sheet
            ws_kpis = wb.active
            ws_kpis.title = 'KPIs'
            
            kpis = report_data['kpis']
            ws_kpis['A1'] = 'RAPPORT MAKÈT PEYIZAN'
            ws_kpis['A2'] = f"Période: {kpis['periode_debut']} au {kpis['periode_fin']}"
            
            ws_kpis['A4'] = 'Métrique'
            ws_kpis['B4'] = 'Valeur'
            
            header_fill = PatternFill(start_color='1a6b2f', end_color='1a6b2f', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ['A4', 'B4']:
                ws_kpis[cell].fill = header_fill
                ws_kpis[cell].font = header_font
            
            row = 5
            for key, value in [
                ('Total Commandes', kpis['total_commandes']),
                ('Chiffre d\'affaires (HTG)', f"{kpis['ca_total']:,.0f}"),
                ('Producteurs Actifs', kpis['total_producteurs']),
                ('Alertes Stock', kpis['alertes_stock']),
                ('Paiements en Attente', kpis['paiements_en_attente']),
            ]:
                ws_kpis[f'A{row}'] = key
                ws_kpis[f'B{row}'] = value
                row += 1
            
            # Daily Sales Sheet
            ws_daily = wb.create_sheet('Ventes Journalières')
            ws_daily['A1'] = 'Date'
            ws_daily['B1'] = 'CA (HTG)'
            
            for cell in ['A1', 'B1']:
                ws_daily[cell].fill = header_fill
                ws_daily[cell].font = header_font
            
            row = 2
            for day in report_data['daily_sales']:
                ws_daily[f'A{row}'] = day['date']
                ws_daily[f'B{row}'] = day['ca']
                row += 1
            
            # Top Products Sheet
            ws_products = wb.create_sheet('Top Produits')
            ws_products['A1'] = 'Produit'
            ws_products['B1'] = 'Quantité'
            ws_products['C1'] = 'CA (HTG)'
            
            for cell in ['A1', 'B1', 'C1']:
                ws_products[cell].fill = header_fill
                ws_products[cell].font = header_font
            
            row = 2
            for p in report_data['top_products']:
                ws_products[f'A{row}'] = p['produit__nom']
                ws_products[f'B{row}'] = p['total_vendu']
                ws_products[f'C{row}'] = p['ca']
                row += 1
            
            # Adjust column widths
            for ws in [ws_kpis, ws_daily]:
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 20
            ws_products.column_dimensions['A'].width = 30
            ws_products.column_dimensions['B'].width = 20
            ws_products.column_dimensions['C'].width = 20
            
            # Save to buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"rapport_maket_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            return HttpResponseServerError(f'Erreur lors de la génération du fichier Excel: {str(e)}')

